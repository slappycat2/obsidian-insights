"""Build a data dictionary of every name v_chk defines, as an .xlsx workbook.

A development aid, not part of the package: ``tools/`` sits outside ``src/``,
so hatchling never sees it and the wheel is unaffected.

    uv run python tools/var_dictionary.py                    # -> tools/_out/
    uv run python tools/var_dictionary.py "F:/Documents/v_chk variables.xlsx"

Three passes, all in this file:

1. **Extract.** Walk main.py and every module in src/vault_check with ``ast``,
   recording each name assigned at module, class or instance level. ast rather
   than regex, so assignment targets and annotations are read structurally and
   a name inside a string or comment is never mistaken for a definition.
2. **Enrich.** Attach the code's own words -- the trailing comment on the
   definition line, and any comment block directly above it.
3. **Build.** Render six sheets, taking purpose text from ``var_purposes.py``.

Locals inside functions are out of scope, and so is ``tests/``.

Two things the output states about itself, and should keep stating:

* **Every purpose says where it came from**, colour-coded on the sheet. 400-odd
  descriptions cannot all be equally well founded, and presenting them as if
  they were would be worse than useless.
* **Occurrence counts are per NAME, not per owner.** They are whole-word text
  matches across the package, so the eight classes that each hold a ``tab_def``
  all report the same total. Read them as a rough signal, not a call graph.
"""
import ast
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

TOOLS = Path(__file__).resolve().parent
ROOT = TOOLS.parent
sys.path.insert(0, str(TOOLS))
from var_purposes import resolve                              # noqa: E402

SRC = [ROOT / "main.py"] + sorted((ROOT / "src" / "vault_check").glob("*.py"))

# Literal on the right-hand side -> content type. Anything not listed leaves the
# type blank, which the workbook shows as a dash: honest about not knowing.
CALL_TYPES = {
    "dict": "dict", "list": "list", "set": "set", "str": "str", "int": "int",
    "float": "float", "bool": "bool", "tuple": "tuple", "defaultdict": "dict",
    "Path": "Path", "compile": "re.Pattern", "now": "datetime",
    "strftime": "str", "today": "date", "StringVar": "tk.StringVar",
    "BooleanVar": "tk.BooleanVar", "IntVar": "tk.IntVar",
    "PhotoImage": "PhotoImage", "Tk": "tk.Tk", "Frame": "ttk.Frame",
    "Label": "ttk.Label", "Button": "ttk.Button", "Entry": "ttk.Entry",
    "Combobox": "ttk.Combobox", "Checkbutton": "ttk.Checkbutton",
    "Spinbox": "ttk.Spinbox", "LabelFrame": "ttk.LabelFrame",
    "Workbook": "openpyxl.Workbook", "safe_load": "dict|list",
    "getLogger": "Logger", "rglob": "generator", "glob": "list",
}


# =========================================================== pass 1: extract ==
def literal_type(node):
    if node is None:
        return ""
    if isinstance(node, ast.Constant):
        return {bool: "bool", int: "int", float: "float", str: "str",
                type(None): "None"}.get(type(node.value), type(node.value).__name__)
    if isinstance(node, (ast.Dict, ast.DictComp)):
        return "dict"
    if isinstance(node, (ast.List, ast.ListComp)):
        return "list"
    if isinstance(node, (ast.Set, ast.SetComp)):
        return "set"
    if isinstance(node, ast.Tuple):
        return "tuple"
    if isinstance(node, ast.JoinedStr):
        return "str"
    if isinstance(node, ast.Call):
        f = node.func
        return CALL_TYPES.get(getattr(f, "id", None) or getattr(f, "attr", None) or "", "")
    if isinstance(node, ast.UnaryOp):
        return literal_type(node.operand)
    if isinstance(node, ast.BinOp):
        return literal_type(node.left) or literal_type(node.right)
    return ""


def unparse(node):
    if node is None:
        return ""
    try:
        s = " ".join(ast.unparse(node).split())
    except Exception:
        return ""
    return s if len(s) <= 90 else s[:87] + "..."


def extract():
    records, sources = {}, {}

    def add(key, **kw):
        """Merge repeat assignments. 'None' is a weak type: an attribute declared
        None in __init__ and given a real value later reports the real one."""
        seen = kw.pop("ctype", "")
        if key in records:
            rec = records[key]
            rec["defs"].append(kw.pop("def_site"))
            if seen:
                rec["seen_types"].append(seen)
            for k, v in kw.items():
                if v and not rec.get(k):
                    rec[k] = v
            return
        kw["defs"] = [kw.pop("def_site")]
        kw["seen_types"] = [seen] if seen else []
        records[key] = kw

    for path in SRC:
        text = path.read_text(encoding="utf-8")
        sources[path.name] = text

        class Visitor(ast.NodeVisitor):
            def __init__(self):
                self.cls = self.fn = None

            def visit_ClassDef(self, node):
                prev, self.cls = self.cls, node.name
                self.generic_visit(node)
                self.cls = prev

            def visit_FunctionDef(self, node):
                prev, self.fn = self.fn, node.name
                self.generic_visit(node)
                self.fn = prev

            def visit_AnnAssign(self, node):
                site = f"{path.name}:{node.lineno}"
                ann = unparse(node.annotation)
                t = node.target
                if isinstance(t, ast.Name) and self.cls and not self.fn:
                    add(f"{self.cls}.{t.id}", name=t.id, kind="dataclass field",
                        owner=self.cls, def_site=site,
                        ctype=ann or literal_type(node.value), default=unparse(node.value))
                elif isinstance(t, ast.Attribute) and isinstance(t.value, ast.Name) \
                        and t.value.id == "self":
                    add(f"{self.cls}.{t.attr}", name=t.attr, kind="instance attribute",
                        owner=self.cls or "?", def_site=site,
                        ctype=ann or literal_type(node.value), default=unparse(node.value))
                self.generic_visit(node)

            def visit_Assign(self, node):
                site = f"{path.name}:{node.lineno}"
                ctype, default = literal_type(node.value), unparse(node.value)
                for target in node.targets:
                    for t in ast.walk(target):
                        if isinstance(t, ast.Attribute) and isinstance(t.value, ast.Name) \
                                and t.value.id == "self":
                            add(f"{self.cls}.{t.attr}", name=t.attr,
                                kind="instance attribute", owner=self.cls or "?",
                                def_site=site, ctype=ctype, default=default)
                        elif isinstance(t, ast.Name) and self.cls and not self.fn:
                            add(f"{self.cls}.{t.id}", name=t.id, kind="class attribute",
                                owner=self.cls, def_site=site, ctype=ctype, default=default)
                        elif isinstance(t, ast.Name) and not self.cls and not self.fn:
                            add(f"<module {path.name}>.{t.id}", name=t.id,
                                kind="module constant", owner=path.name,
                                def_site=site, ctype=ctype, default=default)
                self.generic_visit(node)

        Visitor().visit(ast.parse(text, filename=str(path)))

    settle_types(records)
    count_uses(records, sources)
    attach_comments(records, sources)
    return records


def settle_types(records):
    """An annotated dataclass field wins. Otherwise take the most common concrete
    literal type over a bare None, and fall back to the SysConfig field of the
    same name -- ExcelExporter.sys_pn_wbs is the same string SysConfig declares."""
    ann_by_name = {}
    for rec in records.values():
        if rec["kind"] == "dataclass field" and rec.get("seen_types"):
            t = rec["seen_types"][0]
            if t and t != "None":
                ann_by_name.setdefault(rec["name"], t)

    for rec in records.values():
        seen = [t for t in rec.get("seen_types", []) if t]
        concrete = [t for t in seen if t != "None"]
        if concrete:
            rec["ctype"], rec["ctype_note"] = max(set(concrete), key=concrete.count), ""
        elif seen:
            rec["ctype"] = "None at init"
            rec["ctype_note"] = "assigned None in __init__; real type set later"
        else:
            rec["ctype"] = rec["ctype_note"] = ""
        if not rec["ctype"] or rec["ctype"] == "None at init":
            borrowed = ann_by_name.get(rec["name"])
            if borrowed:
                rec["ctype"] = borrowed
                rec["ctype_note"] = "inferred from the SysConfig field of the same name"


def count_uses(records, sources):
    for rec in records.values():
        pat = re.compile(rf"\b{re.escape(rec['name'])}\b")
        uses = {f: len(pat.findall(t)) for f, t in sources.items()}
        rec["uses"] = {f: n for f, n in uses.items() if n}
        rec["total_uses"] = sum(rec["uses"].values())
        rec["files"] = len(rec["uses"])


# ============================================================ pass 2: enrich ==
def attach_comments(records, sources):
    """The author's own words beat anything invented, so lift them where present."""
    lines = {n: t.splitlines() for n, t in sources.items()}

    def clean(c):
        return re.sub(r"^:\w+\s+\w+:\s*", "", c.strip().lstrip("#").strip())

    for rec in records.values():
        inline, above = "", []
        for site in rec["defs"]:
            fname, lineno = site.rsplit(":", 1)
            src = lines.get(fname, [])
            i = int(lineno) - 1
            if not (0 <= i < len(src)):
                continue
            if not inline:
                m = re.search(r"#\s*(.+)$", src[i])
                if m and not src[i].strip().startswith("#"):
                    inline = clean(m.group(1))
            if not above:
                j, block = i - 1, []
                while j >= 0 and src[j].strip().startswith("#"):
                    block.insert(0, clean(src[j]))
                    j -= 1
                above = [b for b in block if b and not set(b) <= set("-=* ")]
        rec["comment_inline"] = inline
        rec["comment_above"] = " ".join(above)[:400]


# ============================================================= pass 3: build ==
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(size=16, bold=True, color="1F3864")
SUB_FONT = Font(size=10, italic=True, color="595959")
MONO = Font(name="Consolas", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SOURCE_FILL = {
    "authored":       PatternFill("solid", fgColor="E2EFDA"),
    "source comment": PatternFill("solid", fgColor="DDEBF7"),
    "pattern":        PatternFill("solid", fgColor="FFF2CC"),
    "derived":        PatternFill("solid", fgColor="F2F2F2"),
}

CTOT = [
    ("ctot[0]",  "00-Total MD Files in Vault",       "Every .md file rglob() returned."),
    ("ctot[1]",  "01-Templates Processed",           "Files under the Templater folder. Counted here, NOT in ctot[3]."),
    ("ctot[2]",  "02-MD Files in skip_rel_str",      "Files skipped because they sit under an ignored directory."),
    ("ctot[3]",  "03-Files Processed/Dups Teste",    "Files actually harvested and duplicate-tested."),
    ("ctot[4]",  "04-Known Nested Tags Files Found", "NestedDictionary resets -- nested YAML blocks routed to a plugin bucket."),
    ("ctot[5]",  "05-Frontmatter YAML Files",        "Files that had a frontmatter block."),
    ("ctot[6]",  "06-Inline YAML Files",             "Files carrying inline key:: value properties in the body."),
    ("ctot[7]",  "07-upd_obs_files",                 "Calls to upd_obs_files() -- per-file records written."),
    ("ctot[8]",  "08-upd_obs_nests",                 "Calls to upd_obs_nests() -- nested/plugin records written."),
    ("ctot[9]",  "09-upd_obs_props",                 "Calls to upd_obs_props() -- property records written."),
    ("ctot[10]", "10-Files With No Frontmatter",     "Files with no frontmatter at all. Normal in Obsidian, not an error."),
    ("ctot[11]", "11-Max Props",                     "Most links seen for any one property value. Caps the FileNN columns on the Values tab, with link_lim_vals."),
    ("ctot[12]", "12-Max Tags",                      "Most links seen for any one tag. Caps the FileNN columns on the Tags tab, with link_lim_tags."),
    ("ctot[13]", "13-Empty Notes (whitespace only)", "Notes whose raw text is whitespace only. Added in v0.4.0. A SUBSET of slot 10, since an empty note has no frontmatter either."),
]

SINKS = [
    ("obs_props", "{key: {value: [filepath, ...]}}",   "Properties, Values", "Frontmatter and inline properties."),
    ("obs_atags", "{key: {value: [filepath, ...]}}",   "Tags",               "Tags found in note bodies."),
    ("obs_xyaml", "{code: {detail: [filepath, ...]}}", "Possible Issues",    "Bad frontmatter, classified BadY / NoFm / MtFm / ErrY / NonD."),
    ("obs_dupfn", "{filename: {path: [filepath, ...]}}", "Duplicates",       "Filenames occurring more than once."),
    ("obs_files", "{filepath|F-or-I: {key: [values]}}", "Files",             "Per-file view. The only place original key casing survives."),
    ("obs_tmplt", "{key: {value: [filepath, ...]}}",   "Templates",          "Templater files. Deliberately kept out of every other sink."),
    ("obs_codes", "{key: {value: [filepath, ...]}}",   "Code",               "Code blocks by language/signature."),
    ("obs_nests", "{plugin_id|filepath: {key: [values]}}", "Nested",         "Nested YAML, treated as plugin-managed data."),
    ("obs_plugs", "{key: {value: [filepath, ...]}}",   "Plugins",            "Installed plugins from manifest.json + community-plugins.json."),
    ("obs_empty", "[filepath, ...]",                   "Possible Issues",    "Whitespace-only notes. The one flat list; ExcelExporter turns it into a set."),
]

CFG_KEYS = """sys_id sys_ver sys_dir_sys sys_dir_dat sys_dir_bat sys_dir_wbs sys_dir_log sys_dir_img
sys_pn_cfg sys_pn_wb_exec sys_pn_batch sys_pn_wbs sys_tab_seq sys_cfg_os cur_vlts sys_vlts sys_pn_lg2
sys_pn_lg3 sys_pn_ico sys_pn_bnr sys_pn_a51 sys_splash_bg vault_name vault_id dir_vault dir_templates
skip_rel_str skip_abs_lst dirs_dot ctot bool_shw_notes bool_rel_paths bool_summ_rows bool_unused_1
bool_unused_2 bool_unused_3 link_lim_vals link_lim_tags v_chk_date""".split()

ON_SCREEN = {"vault_name", "skip_rel_str", "sys_pn_wb_exec", "bool_shw_notes",
             "bool_rel_paths", "bool_unused_1", "bool_unused_2",
             "link_lim_vals", "link_lim_tags"}


def build(records, out_path):
    wb = Workbook()

    def sheet(title, headings, widths, rows, blurb, table_name=None):
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        ws["A2"] = blurb
        ws["A2"].font = SUB_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headings))
        ws.row_dimensions[2].height = 30
        ws["A2"].alignment = WRAP

        for c, h in enumerate(headings, 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BOX
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[4].height = 22

        for r, row in enumerate(rows, 5):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = BOX
                cell.alignment = WRAP if widths[c - 1] > 30 else TOP
                if headings[c - 1] in ("Name", "Defined at", "Type", "Default"):
                    cell.font = MONO
                if headings[c - 1] == "Purpose from":
                    cell.fill = SOURCE_FILL.get(val, SOURCE_FILL["derived"])

        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

        if rows and table_name:
            ref = f"A4:{get_column_letter(len(headings))}{4 + len(rows)}"
            t = Table(displayName=table_name, ref=ref)
            t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
            ws.add_table(t)

        ws.freeze_panes = ws.cell(row=5, column=2)
        return ws

    rows, counts = [], defaultdict(int)
    for key, rec in sorted(records.items(), key=lambda kv: (kv[1]["owner"], kv[1]["name"])):
        purpose, source = resolve(key, rec)
        counts[source] += 1
        ctype = rec["ctype"] or "\u2014"
        if rec.get("ctype_note", "").startswith("inferred"):
            ctype += " (inferred)"
        rows.append([
            rec["name"], rec["kind"], rec["owner"], ctype, rec["default"] or "\u2014",
            "; ".join(rec["defs"][:3]),
            ", ".join(f"{f} ({n})" for f, n in sorted(rec["uses"].items(), key=lambda kv: -kv[1])),
            rec["files"], rec["total_uses"], purpose, source,
        ])

    sheet("All variables",
          ["Name", "Kind", "Defined on", "Type", "Default / initial value", "Defined at",
           "Used in (file: occurrences)", "Files", "Uses", "Purpose", "Purpose from"],
          [26, 18, 20, 16, 30, 26, 46, 7, 7, 78, 15], rows,
          "Every name assigned at module, class or instance level across main.py and src/vault_check, "
          "extracted with Python's ast module. A type marked (inferred) was borrowed from the SysConfig "
          "field of the same name; 'None at init' means the attribute starts as None and is given its "
          "real value later; a dash means the type depends on what is assigned at run time. Occurrence "
          "counts are whole-word matches and include the definition. Colour on the last column shows "
          "how much to trust the Purpose: green = written after reading the code, blue = the code's own "
          "comment, amber = a rule applied to a family of names, grey = no description available.",
          "tbl_all")

    cfg_rows = []
    for key, rec in sorted(records.items(), key=lambda kv: kv[1]["name"]):
        if rec["owner"] != "SysConfig" or rec["kind"] != "dataclass field":
            continue
        purpose, _ = resolve(key, rec)
        cfg_rows.append([rec["name"], rec["ctype"] or "\u2014", rec["default"] or "\u2014",
                         "yes" if rec["name"] in CFG_KEYS else "no",
                         "yes" if rec["name"] in ON_SCREEN else "no",
                         rec["files"], purpose])

    sheet("Settings (SysConfig)",
          ["Name", "Type", "Default", "In packed sys_cfg?", "On setup screen?",
           "Files touching it", "Purpose"],
          [24, 12, 34, 17, 16, 13, 95], cfg_rows,
          "The SysConfig dataclass -- what CONFIG.yaml stores. The fourth column matters: downstream "
          "stages read the PACKED sys_cfg dict, never these attributes, so any attribute changed after "
          "load_config() must be followed by cfg_pack() or the change is silently ignored. A 'no' there "
          "means the field is local to SysConfig and never reaches the pipeline.",
          "tbl_cfg")

    sheet("Counters (ctot)", ["Slot", "Area51 label", "Counts", "Type"], [11, 36, 92, 10],
          [[a, b, c, "int"] for a, b, c in CTOT],
          "sys_cfg['ctot'] -- a list of integers incremented through v_chk_build.py and rendered on the "
          "Area51 tab, not the Summary tab. Its length is CTOT_SLOTS in src/vault_check/__init__.py, the "
          "one place it is stated. Adding a slot means bumping CTOT_SLOTS, appending to ctot_descs AND "
          "adding the matching f-tot-NN / x-tot-NN cell pair in DefAr51.tab_cd_fixed_summ -- a "
          "description with no cell is simply never rendered.",
          "tbl_ctot")

    sheet("Data sinks", ["Name", "Shape", "Tab that consumes it", "Holds", "Type"],
          [14, 36, 20, 72, 10],
          [[a, b, c, d, "list" if a == "obs_empty" else "dict"] for a, b, c, d in SINKS],
          "The harvest. VaultHealthCheck accumulates into these, they travel inside wb_def['wb_data'], "
          "and each tab reads exactly one via its data_src. A tab whose sink is empty is dropped rather "
          "than rendered, which is why a vault with no Templater folder legitimately produces 10 sheets, "
          "not 12.",
          "tbl_sinks")

    once = []
    for key, rec in sorted(records.items(), key=lambda kv: (kv[1]["owner"], kv[1]["name"])):
        if rec["total_uses"] == 1:
            purpose, _ = resolve(key, rec)
            once.append([rec["name"], rec["owner"], rec["ctype"] or "\u2014",
                         rec["defs"][0], purpose])

    sheet("Written once", ["Name", "Defined on", "Type", "Defined at", "Purpose"],
          [26, 20, 16, 26, 100], once,
          "Names that occur exactly once in the whole package -- the definition itself, and no read "
          "anywhere. Candidates for removal, or a sign that something meant to use them never got wired "
          "up. Not proof of either: a name reached only through getattr(), a YAML key or a Tk trace will "
          "look unused here. v0.4.1 removed a batch of exactly this kind of thing, including a "
          "self.obs_plugs that never existed because the line above it overwrote it.",
          "tbl_once")

    write_readme(wb, len(rows), len(once), counts)
    wb.move_sheet("Read me", offset=-6)
    wb.save(out_path)
    return counts


def write_readme(wb, n_rows, n_once, counts):
    ws = wb["Sheet"]
    ws.title = "Read me"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "v_chk -- variable dictionary"
    ws["A1"].font = Font(size=20, bold=True, color="1F3864")
    ws.column_dimensions["A"].width = 118

    notes = [
        ("", ""),
        (f"Generated {datetime.now():%Y-%m-%d %H:%M} by tools/var_dictionary.py.", "sub"),
        ("", ""),
        ("What is in here", "h"),
        (f"Every name assigned at module, class or instance level in main.py and the modules of "
         f"src/vault_check -- {n_rows} of them. Locals inside functions are not included; neither "
         f"are the tests.", ""),
        ("", ""),
        ("How it was built", "h"),
        ("Names, types and defaults come from Python's ast module, so assignment targets and "
         "annotations are read structurally rather than by matching text.", ""),
        ("Occurrence counts are whole-word text matches across the package, and include the definition "
         "itself. Read them as a rough signal, not a call graph.", ""),
        ("One consequence to keep in mind: the count is per NAME, not per owner. Eight classes each "
         "hold a tab_def, so all eight rows report the same total -- that is every occurrence of the "
         "word tab_def anywhere in the package, not that many uses of one attribute. Short or common "
         "names (name, title, version, id) are inflated the same way. The 'Defined on' column is what "
         "distinguishes the rows; the count does not.", ""),
        (f"A count of 1 means the name is written once and never read again. There are {n_once} of "
         f"those, listed on the 'Written once' sheet.", ""),
        ("", ""),
        ("How much to trust the Purpose column", "h"),
        ("The last column of 'All variables' says where each description came from, and the cell is "
         "coloured to match. This matters: hundreds of descriptions cannot all be equally well founded, "
         "and it would be worse to present them as if they were.", ""),
        ("    authored       -- written after reading the code and CLAUDE.md. Trust these.", "mono"),
        ("    source comment -- the code's own words, lifted from a comment at the definition.", "mono"),
        ("    pattern        -- a rule applied to a family of names (every *_var on the setup screen, "
         "every f_uniq_* formula). Right about the family; may miss what is specific to one member.", "mono"),
        ("    derived        -- nothing was available. The cell says so and points at the line.", "mono"),
        ("To improve a row, move it up to AUTHORED in tools/var_purposes.py and re-run.", ""),
        ("", ""),
        ("The other sheets", "h"),
        ("Settings (SysConfig)   the fields CONFIG.yaml stores, and whether each reaches the pipeline", "mono"),
        ("Counters (ctot)        the Area51 counter slots and what increments each", "mono"),
        ("Data sinks             the harvest structures, their shapes, and which tab consumes each", "mono"),
        ("Written once           the names with no read anywhere in the package", "mono"),
        ("", ""),
        ("One more thing", "h"),
        ("A name assigned in more than one place is one row, listing the first three definition sites. "
         "Same-named attributes on different classes are separate rows.", ""),
        ("This is a snapshot with no link back to the source, so it goes stale the moment the code "
         "changes. Re-run tools/var_dictionary.py rather than editing it by hand.", ""),
    ]

    r = 3
    for text, style in notes:
        c = ws.cell(row=r, column=1, value=text)
        c.font = {"h": Font(size=12, bold=True, color="1F3864"),
                  "sub": SUB_FONT,
                  "mono": Font(name="Consolas", size=10)}.get(style, Font(size=11))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if len(text) > 110:
            ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Purpose sources").font = Font(size=12, bold=True, color="1F3864")
    r += 1
    for src, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        cell = ws.cell(row=r, column=1, value=f"    {n:>4}   {src}")
        cell.font = Font(name="Consolas", size=10)
        cell.fill = SOURCE_FILL[src]
        r += 1


def main():
    if len(sys.argv) > 1:
        out = Path(sys.argv[1])
        out.parent.mkdir(parents=True, exist_ok=True)
    else:
        out_dir = TOOLS / "_out"
        out_dir.mkdir(parents=True, exist_ok=True)
        out = out_dir / "v_chk variables.xlsx"

    records = extract()
    counts = build(records, out)

    print(f"{len(records)} names from {len(SRC)} files")
    for src, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        print(f"  {n:>4}  purpose: {src}")
    print(f"wrote {out}")

    if "--json" in sys.argv:
        side = out.with_suffix(".json")
        side.write_text(json.dumps(records, indent=1), encoding="utf-8")
        print(f"wrote {side}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
