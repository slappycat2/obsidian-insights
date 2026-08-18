"""Render the analysed vault data into a formatted .xlsx workbook.

Consumes the per-tab cell definitions built by v_chk_wb_tabs and writes tables,
conditional formatting, obsidian:// hyperlinks and images through openpyxl.

Outstanding bugs and enhancements live in docs/BACKLOG.md, and on GitHub once
the repository is published -- not in comments here.
"""

import sys
import os
import time
import copy
import re
import urllib.parse

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule  # , ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import tkinter as tk
from tkinter import messagebox

from vault_check.v_chk_colors import Colors
from vault_check.v_chk_plugin_man import PluginMan
from vault_check.v_chk_logger import logger


FALLBACK_FONT = "Arial"


class ExcelExporter:
    def __init__(self, wbd_obj):
        self.tab_def = {}
        self.wb_tabs_open = {}
        self.v_chk_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.TAG_TOT_COLS = 20  # Set this to the maximum columns to be added to table (w/o the step)
        self.TAG_BEG_COL = 7  # set the to the first column, ie, where to start
        self.PROP_TOT_COLS = 13  # Set this to the maximum columns to be added to table (w/o the step)
        self.PROP_BEG_COL = 9  # set the to the first column, ie, where to start
        self.COL_STEP = 2  # This should be the number of cols in each loop
        self.TABLE_GROUPINGS = False
        self.next_cell_col = 0
        self.last_cell_row = 0
        self.colors = Colors()

        self.tab_id_sub_key = ''
        self.tabs_built = {}
        self.wb_tabs_done = {}

        # wbd_obj = WbDataDef()
        self.xyml_descs = wbd_obj.xyml_descs

        self.wb_def = wbd_obj.read_wb_data()
        self.sys_cfg = self.wb_def.get('sys_cfg', {})

        self.sys_tab_seq = self.sys_cfg['sys_tab_seq']   # ['summ', 'pros', 'tags', 'dups', 'xyml']  # ['summ', 'pros', 'tags', 'dups']
        self.dir_vault = self.sys_cfg['dir_vault']
        self.vault_id = self.sys_cfg['vault_id']
        self.sys_pn_wbs = self.sys_cfg['sys_pn_wbs']
        logger.info(f"initializing workbook {self.sys_pn_wbs}...")

        self.sys_pn_batch = self.sys_cfg['sys_pn_batch']
        self.sys_pn_wb_exec = self.sys_cfg['sys_pn_wb_exec']

        self.wb_data = self.wb_def.get('wb_data', {})
        # Only obs_empty is mirrored onto an attribute, because the Xyml row
        # builder needs a per-row membership test. Every tab reads its own data
        # through wb_def['wb_data'][data_src] instead, so the other nine sinks
        # used to be copied here and never read again -- and the copy had drifted:
        # obs_nests was assigned twice, the second time from obs_plugs, so
        # self.obs_plugs was never created at all. Do not reinstate them.
        #
        # A list on the wire so it stays yaml.dump-able; a set here for the test.
        self.obs_empty = set(self.wb_data.get('obs_empty', []))

        # Compiled once: export_cell() runs this over every string cell in the
        # workbook, and re.compile() on each was the only reason it was a str.
        self.rgx_noTZdatePattern = re.compile(
            r"([0-9]{4})[-\/]([0-1]?[0-9]{1})[-\/]([0-3])?([0-9]{1})(\s+)([0-9]{2}:[0-9]{2}:[0-9]{2})(.*)")
        self.rgx_noTZdateReplace = r"\1-\2-\3\4 \6"

        # self.code_q_types = ['TABLE', 'LIST', 'TASK', 'CALENDAR']
        self.plugin_lib = PluginMan(self.dir_vault)

        logger.debug(f"ExcelExport - cfg.sys_pn_wb_exec: {self.sys_pn_wb_exec}")

        self.exl_file = Path(self.sys_pn_wb_exec)

    def export(self):
    # =================================================================================

        logger.debug(f"ExcelExport.export - sys_pn_wb_exec: {self.sys_pn_wb_exec}")

        # Create the workbook instance
        wb = openpyxl.Workbook()

        wb = self.initialize_all_tabs(wb)

        for tab_id in self.sys_tab_seq:
            logger.debug(f"Exporting tab: {tab_id}")

            self.tab_def = self.wb_def['wb_tabs'][tab_id]
            # self.tab_def = self.wb_tabs_open[tab_id]
            self.tabs_built[tab_id] = self.tab_def
            if tab_id != 'ar51':
                self.wb_tabs_done[tab_id] = self.export_tab()
            
        self.export_area51()

        logger.debug(f"Building workbook completed successfully.")

        self.save_workbook(wb)

    def export_area51(self) -> None:
        tab_id = 'ar51'
        tab = self.wb_tabs_open[tab_id]
        err_txt = self.colors.err_txt

        tab_def = self.wb_def['wb_tabs'][tab_id]

        # ========================================================================
        # export Totals Grid, both headers and formulas for totals
        # ========================================================================
        val = ''
        # row_idx = 19
        tab = self.export_grid(tab, 1)
        # ========================================================================
        # export cfg
        # ========================================================================
        cfg_cd_def = self.tab_def['cfg-dump']
        col_idx = cfg_cd_def[0]
        row_idx = cfg_cd_def[1]
        col_sav = col_idx

        for key, value in self.sys_cfg.items():
            if isinstance(value, (list, tuple)):
                val = ', '.join([str(item) for item in value])
                val = f"[{val}]"
            elif isinstance(value, dict):
                val = str(value)
            else:
                val = value

            _, cell = self.export_cell(tab, cfg_cd_def, key, row_idx)
            cfg_cd_def[0] = 0
            _, cell = self.export_cell(tab, cfg_cd_def, val, row_idx)
            cfg_cd_def[0] = col_sav
            row_idx += 1

        tab.conditional_formatting.add('E21:E28', CellIsRule(
            operator='notEqual', formula=['0'], stopIfTrue=False, font=Font(color=err_txt, bold=True, italic=True)))

    def export_tab(self):
        tab_id = self.tab_def['tab_id']
        tab_def = self.tab_def
        tab = self.wb_tabs_open[tab_id]

        tab_table_style = tab_def['tab_table_style']
        tbl_hdr_row = tab_def['tbl_hdr_row']
        tbl_beg_col = tab_def['tbl_beg_col']
        tab_table_links_cols = tab_def['tab_table_links_cols']
        tab_tots_isVisible_col = tab_def['tab_tots_isVisible_col']
        tbl_name = tab_def['tbl_name']
        tab_cd_table_hdr = tab_def['tab_cd_table_hdr']
        tab_cd_table_dtl = tab_def['tab_cd_table_dtl']
        hdr_IsVis        = tab_def['hdr_IsVis']
        showGridLines    = self.tab_def['showGridLines']
        data_src           = tab_def['data_src']

        logger.debug(f"Exporting Tab: {tab_id}") #  Set a breakpount on this line

        # Create the tab, or rename Sheet 1, in the Summaries case...
        # if tab_id == 'summ':
        #     tab = wb.active
        #     tab.title = tab_name
        # else:
        #     tab = wb.create_sheet(title=tab_name)  # Excel allows max 31 characters in tab names

        # tab.sheet_properties.tabColor = tab_color
        # tab.sheet_view.showGridLines = showGridLines

        # TAB HEADINGS export the Tabs Main Title in a large font
        temp_def = self.tab_def['tab_cd_title_def']
        row_idx = temp_def[1]
        _, cell = self.export_cell(tab, temp_def, '', row_idx)

        # help_txt, subtitles, notes, etc.
        if 'tab_help_txt' in tab_def:
            for hlp_key, hlp_texts in tab_def['tab_help_txt'].items():
                texts_def = tab_def[f'tab_cd_{hlp_key}_def']
                for row_idx, text_line in enumerate(hlp_texts, start=texts_def[1]):
                    _, cell = self.export_cell(tab, texts_def, text_line, row_idx)

        # TABLE-Hdr export the main body table - column headers
        for hdr_key, hdr_col_def_list in tab_cd_table_hdr.items():
            # unpack header definitions
            col_idx, cell = self.export_cell(tab, hdr_col_def_list, '', tbl_hdr_row)
            logger.debug(f"Processing [Header]:{hdr_key}")
            logger.debug(f" {tab_id}  row:{tbl_hdr_row}  col: {col_idx}  list: {hdr_col_def_list}")

        if tbl_beg_col != 0:
            tab[f"{self.xl_a_col(tbl_beg_col + 1)}{tbl_hdr_row + 1}"] = "Nothing Found."

        # TABLE-Dtl Begin Detail
        row_idx = tbl_hdr_row + 1
        p_v_Index_p_count, p_v_Index_v_count = 0, 0
        beg_prop_group = row_idx + 1
        last_prop_name = ""

        # ==============================================================================
        # TABLE-Dtl: Gathering data--Loop each prop_name, getting the list of all values
        # ==============================================================================
        # First, get the tabs data_src
        tab_src_data = self.wb_def['wb_data']
        for src in data_src:
            tab_src_data = tab_src_data[src]

        for prop_name, values_dict in tab_src_data.items():

            if tab_id == 'summ':
                continue

            # so for dups, values_dict will be all pathnames using this filename
            prop_name = self.xl_clean_cell(prop_name)
            p_v_Index_p_count += 1
            p_v_Index_v_count = 0
            if tab_id == "vals" and last_prop_name == "":
                last_prop_name = prop_name

            # sorted_values_dict = {}
            # sorted_values_list = sorted(values_dict.keys())    # this creates a list of keys, not a dict
            # for value_key in sorted_values_list:       # so, now we need to rebuild the dictionary
            #     sorted_values_dict[value_key] = values_dict[value_key]

            # TABLE-Dtl Now, build vals[], by looping each value,  getting a list of Files-Where-Used
            # for value_item, value_files_list in values_dict.items():
            for value_item, value_files_list in values_dict.items():
                if tab_id == 'dups' and len(value_files_list) == 1:
                    continue
                # reset tab_def to default, in case any changes were made on the prev row
                this_row_dtl = copy.deepcopy(tab_cd_table_dtl)
                # First, define what the values are going to be for this row
                value_item_count = len(value_files_list)
                vals = []
                p_v_Index_v_count += 1

                # TABLE Dtl - Set up First (Fixed) Columns Values
                if tab_id == "pros":
                    # The Properties tab is one row per property, so only the
                    # first value of each property emits, and it counts the
                    # property as a whole.
                    #
                    # The previous scheme emitted at the *boundary* between two
                    # properties instead, carrying running counters that were
                    # reset mid-property. Every row was therefore named after
                    # the preceding property, its file count dropped that
                    # property's own first value and borrowed the next
                    # property's, the first row was a phantom holding no
                    # property at all, and the last property in the vault never
                    # got a row.
                    if p_v_Index_v_count > 1:
                        continue

                    # Distinct files, not the sum of per-value file lists: one
                    # note listing two values of the same property is one file.
                    distinct_files = {file
                                      for files in values_dict.values()
                                      for file in files}
                    vals = [int((row_idx - tbl_hdr_row))
                        , prop_name
                        , len(values_dict)
                        , len(distinct_files)
                        ]
                elif tab_id == "vals":
                    vals = [int((row_idx - tbl_hdr_row))
                        , prop_name
                        , value_item
                        , value_item_count
                        , f"{p_v_Index_p_count:03d}-{p_v_Index_v_count:05d}"
                            ]
                elif tab_id == "tags":
                    vals = [int((row_idx - tbl_hdr_row))
                        , value_item
                        , value_item_count
                            ]
                elif tab_id == "dups":
                    vals = [int((row_idx - tbl_hdr_row))
                        , self.obs_hyperlink(Path(value_item).name)
                        , value_item_count
                            ]
                elif tab_id == "xyml":
                    # For an xyml row, value_item is the note's path. An empty note
                    # has nothing for the Files tab to match, so the usual lookup
                    # would just read "" -- say why instead.
                    if value_item in self.obs_empty:
                        fm_okay = '(empty file)'
                    else:
                        fm_okay = '=IFERROR(IF(VLOOKUP(tbl_xyml[Notes],tbl_file[Notes],1,FALSE)=tbl_xyml[Notes],TRUE,FALSE),"")'

                    vals = [int((row_idx - tbl_hdr_row))
                            , self.obs_hyperlink(Path(value_item).name)
                            , fm_okay
                            , self.xyml_descs[prop_name][0]
                            ]
                elif tab_id == "file":
                    file_nm, loc = prop_name.split("|")
                    if loc == 'F':
                        loc = ''
                    else:
                        loc = '(Inline)'

                    act_prop = value_files_list[0]
                    if act_prop == value_item:
                        act_prop = ""
                    prop_vals = " | ".join(map(str, value_files_list[1:]))

                    vals = [int((row_idx - tbl_hdr_row))
                            , self.obs_hyperlink(Path(file_nm).name)
                            , ' '
                            , loc
                            , value_item
                            , act_prop
                            , len(value_files_list) - 1
                            , prop_vals
                            ]
                elif tab_id == "code":
                    file_nm = prop_name
                    cb_sig = value_item
                    plugin_id = self.plugin_lib.get_name(cb_sig)

                    vals = [int((row_idx - tbl_hdr_row))
                            , self.obs_hyperlink(Path(file_nm).name)
                            , plugin_id
                            , cb_sig
                            , len(value_files_list)
                            ]
                elif tab_id == 'nest':
                    plugin_id, file_nm = prop_name.split("|")
                    vtot = len(value_files_list)
                    prop_vals = " | ".join(map(str, value_files_list[1:]))
                    vals = [int((row_idx - tbl_hdr_row))
                            , plugin_id
                            , self.obs_hyperlink(Path(file_nm).name)
                            , value_item
                            , vtot
                            , prop_vals
                            ]
                elif tab_id == 'plug':
                    # , 0 'id':                     # - cmdr (not used; overwritten with RowId)
                    # , 1 'name':                   # - Commander
                    # , 2 'enable':                 # - true
                    # , 3 'version':                # - 0.5.2
                    # , 4 'minAppVersion':          # - 1.4.0
                    # , 5 'author':                 # - bob, carol, ted & alice
                    # , 6 'authorUrl':              # - https://github.com/phibr0
                    # , 7 'isDesktopOnly':          # - false
                    # , 8 'Description':            # - Customize your workspace by adding commands
                    # , 9 sig list this plugin uses # []
                    vals = value_files_list
                    vals[0] = int(row_idx - tbl_hdr_row) # overwrite the id, not using it...
                    if "http" in vals[6].lower():
                        vals[6] = self.web_hyperlink(vals[6])
                    if vals[7] is False:
                        vals[7] = ""

                    if vals[2] is False:
                        vals[2] = ""
                        this_row_dtl['name'][8] = True  # set name to red, italic
                        this_row_dtl['name'][5] = self.colors.clr_red

                    if len(vals[9]) > 1:
                        vals[9] = " | ".join(map(str, vals[9]))
                    else:
                        vals[9] = "".join(map(str, vals[9]))

                elif tab_id == "tmpl":
                    vals = [int((row_idx - tbl_hdr_row))
                        , prop_name
                        , value_item
                        , value_item_count
                        , f"{p_v_Index_p_count:03d}-{p_v_Index_v_count:05d}"
                        ]

                # TABLE-Dtl Set up List of Files Used, convert them to obsidian URLs, store in vals[]

                # tab_def["tab_cd_table_links"] =  [9, 0, '', 11, 18, "", "", False, False, 'left'  , '']
                for col_count, file in enumerate(value_files_list, start=1):
                    if col_count > tab_table_links_cols:
                        break

                    obs_link = file
                    if tab_id == "dups":
                        # Need qualified relative path w/o vault path
                        obs_link = self.obs_hyperlink(file.replace(self.dir_vault, ""))
                    else:
                        # Just need the MD filename
                        if isinstance(file, str) and file.endswith('.md') and Path(file).is_file():
                            obs_link = self.obs_hyperlink(Path(file).name)

                    vals = vals + [obs_link, " "]

                # TABLE-Dtl Finally, now we have all the col values in vals, so we can add this row...
                logger.debug(f"Processing [Property]:{prop_name} - [Value]:{value_item} - [Count]: {value_item_count}")

                # ===========================================================================================
                # TABLE-Sub-Dtl Actual start of a new row export, loop thru tab_table columns and add to sheet
                # ===========================================================================================

                for dummy_key, dtl_col_def_list in this_row_dtl.items():
                    col_idx = dtl_col_def_list[0]
                    if col_idx == 0:
                        col_idx = self.next_cell_col
                    # TABLE-Dtl If there is no value in vals for this col
                    if len(vals) <= (col_idx - tbl_beg_col):
                        continue

                    val = vals[col_idx - tbl_beg_col]

                    logger.debug(f"Col: {col_idx}-{dummy_key}")
                    logger.debug(f"Val {col_idx - tbl_beg_col}: {vals[col_idx - tbl_beg_col]}")

                    if col_idx == tab_tots_isVisible_col:
                        val = tab_def['f_isVisible']

                    if isinstance(val,  list):
                        val = " | ".join(map(str, val))

                    _, _ = self.export_cell(tab, dtl_col_def_list, val, row_idx)

                    # if col_idx == 5 and prop_name == last_prop_name and prop_name == "pros":
                    #     cell.font = Font(color=self.tab_def.clr_gry35)

                if self.TABLE_GROUPINGS:
                    if prop_name != last_prop_name and last_prop_name != "":
                        if beg_prop_group != row_idx - 1:
                            tab.row_dimensions.group(beg_prop_group, row_idx - 1, hidden=False)
                            # wgh = openpyxl.worksheet.dimensions.DimensionHolder(tab, reference='index', default_factory=None)
                            # wgh.group(beg_prop_group, row_idx - 1, outline_level=1, hidden=False)
                            beg_prop_group = row_idx + 1

                last_prop_name = prop_name
                row_idx += 1

        # Group last group
        if self.TABLE_GROUPINGS:
            if beg_prop_group != row_idx - 1:
                tab.row_dimensions.group(beg_prop_group, row_idx - 1, hidden=False)

        if self.tab_def['tbl_beg_col']:
            self.format_as_table(tab, tbl_name, tab_table_style, row_idx)

        last_row = row_idx

        tab = self.export_grid(tab, last_row)

        return tab

    def export_grid(self, tab, last_row):
        # ===========================================================================================
        # FIXED CELLS (grid)-These are non-table items that w/"fixed" cell positions on the worksheet
        # ===========================================================================================
        tab_id = self.tab_def['tab_id']
        tab_def = self.tab_def
        tab_tots_isVisible_col = tab_def['tab_tots_isVisible_col']
        tbl_hdr_row = tab_def['tbl_hdr_row']
        hdr_IsVis = tab_def['hdr_IsVis']
        vis_key = 'isVisible'
        tot_table = tab_def['tab_cd_fixed_grid']

        if self.tab_def['tab_has_isVisible_col'] and tab_tots_isVisible_col:
            # STEP 1 - Set IsVisible formula in last table column
            # Set isVisible Column Header

            row_num = tbl_hdr_row
            cell_def = tab_def['tab_cd_fixed_grid'][vis_key]

            col, row, font, sz, w, t_clr, fill_clr, Bold, Ital, Align, val = cell_def
            _, cell = self.export_cell(tab, cell_def, hdr_IsVis, row_num)
            row_num += 1

            # Now, export IsVisible Formula to all cells in the IsVisible column
            cell_def[5] = cell_def[6] = ''   # shut off hdr colors in dtl
            for row_num in range(row_num, last_row):
                _, cell = self.export_cell(tab, cell_def, val, row_num)

        # Now, do the rest of the grid

        val = None    # get rid of the last val
        row_idx = 0     # use the values defined in the grid

        # ========================================================================
        # export Totals Grid, misc totals with "fixed" cell addresses
        # ========================================================================
        for tot_key, tot_col_def_list in tot_table.items():
            if tot_key == vis_key:
                continue

            if tot_key.startswith("img"):
                #   0   1    2  3   4  5      6        7   8       9     10
                # [col,row,font,sz, w,t_clr,fill_clr,Bold,Ital,  Align,  val ] = 11
                a_img = Image(tot_col_def_list[10])
                a_cell = f"{self.xl_a_col(tot_col_def_list[0])}{tot_col_def_list[1]}"
                # add to worksheet and anchor next to cells
                tab.add_image(a_img, a_cell)
                continue

            logger.debug(f"Grid-Tab: '{tab_id} Val:{tot_key}' {tot_col_def_list}")
                # ,{row_idx} val:{val}  def_val: {def_val}")  # Set a breakpount on this line

            _, cell = self.export_cell(tab, tot_col_def_list, val, row_idx)

            val = None

        # ========================================================================
        # export Summary Totals (summ), both headers and formulas for totals
        # ========================================================================
        if 'tab_cd_fixed_summ' in tab_def:
            tot_table = tab_def['tab_cd_fixed_summ']
            for tot_key, tot_col_def_list in tot_table.items():
                logger.debug(f"Grid-summ-Tab: '{tab_id} Val:{tot_key}' {tot_col_def_list}")
                    # ,{row_idx} val:{val}  def_val: {def_val}")  # Set a breakpount on this line

                _, cell = self.export_cell(tab, tot_col_def_list, val, row_idx)

                val = None

        # ========================================================================
        # export borders
        # ========================================================================
        if 'borders' in tab_def:
            for _, brdr_parms in tab_def['borders'].items():
                # 'footer':["C30:J30", "thin", self.colors.clr_blk]
                self.xl_set_border(tab, brdr_parms)

        return tab

    def export_cell(self, tab, col_def_list, val, row_idx):
        # Todo: Change to add the following rules:
        #   If a col_idx = 0 then increment the column by 1, Note that if row=0, row stays the same
        # col,row,font,sz, w,t_clr,f_clr,Bold,Ital,  Align,  val ] = 11
        col_idx, def_row, c_font, c_sz, col_w, txt_clr, fill_clr, bold_bool, ital_bool, align_val, def_val = col_def_list

        logger.debug(f"Exporting Cell (Col,Row): {col_def_list[0]},{row_idx} val:{val}  def_val: {def_val}") #  Set a breakpount on this line

        logger.debug(f"Exporting Cell (Col,Row): {col_def_list[0]},{row_idx} val:{val}  def_val: {def_val}") #  Set a breakpount on this line

        if val is None or val == "":
            val = def_val

        hyperlink = False
        if isinstance(val, str) and val.upper().startswith("=HYPERLINK"):
            hyperlink = True

        if isinstance(val, str) and val.endswith(":"):
            txt_clr = self.colors.clr_red
            ital_bool = True
            bold_bool = True

        if isinstance(val, str) and val in ['alias', 'cssclass', 'tag']:
            txt_clr = self.colors.clr_red
            ital_bool = True
            bold_bool = True
            val += ' (Deprecated in Obsidian 1.4)'

        if row_idx == 0:
            if def_row != 0:
                row_idx = def_row
            else:
                row_idx = self.last_cell_row

        if col_idx == 0:
            col_idx = self.next_cell_col

        if c_sz is None or c_sz == 0:
            c_sz = 10

        if c_font == '' or c_font is None:
            # Every cell that states no preference lands here, including a tab
            # title whose display font is not installed on this machine. Arial
            # is the fallback because an .xlsx names one font with no fallback
            # list: it is present on Windows and macOS, and on Linux fontconfig
            # substitutes metric-compatible Liberation Sans. Naming nothing at
            # all would leave the cell on the workbook default, which is
            # Calibri -- itself an Office font, and so missing in exactly the
            # places this fallback exists to cover.
            c_font = FALLBACK_FONT

        fg_clr = txt_clr
        if fill_clr != "" and fill_clr is not None:
            fg_clr = self.colors.get_txt_clr(fill_clr)

        val = self.xl_clean_cell(val)

        cell = tab.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name=c_font, size=c_sz, bold=bold_bool, italic=ital_bool)
        if align_val == 'wrap':
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        else:
            cell.alignment = Alignment(horizontal=align_val, vertical="center")
        logger.debug(f"Set Cell: at {row_idx}:{col_idx} to:{val}")

        if txt_clr != "" and txt_clr is not None:
            cell.font = Font(name=c_font, size=c_sz, color=txt_clr, bold=bold_bool, italic=ital_bool)

        if hyperlink:
            if txt_clr == "" or txt_clr is None:
                txt_clr = self.tab_def['tab_link_clr']
            cell.font = Font(color=txt_clr, underline='single', bold=True, italic=ital_bool)

        if fill_clr != "":
            cell.fill = PatternFill(start_color=fill_clr, end_color=fill_clr, fill_type="solid", fgColor=fg_clr)

        # col width
        a_col = self.xl_a_col(col_idx)
        if col_w > 0:
            tab.column_dimensions[a_col].width = col_w

        self.last_cell_row = row_idx
        self.next_cell_col = col_idx + 1

        return col_idx, cell

    def initialize_all_tabs(self, wb):
        live_sys_tab_seq = []
        for tab_id in self.sys_tab_seq:
            tab_def = self.wb_def['wb_tabs'][tab_id]
            data_src = tab_def['data_src'][0]
            if len(self.wb_def['wb_data'][data_src]) == 0:
                continue
            else:
                live_sys_tab_seq.append(tab_id)

            tab_name = tab_def['tab_name']
            tab_color = tab_def['tab_color']
            showGridLines = tab_def['showGridLines']

            # Create the tab, or rename Sheet 1, in the Summaries case...
            if tab_id == 'summ':
                tab = wb.active
                # tab.page_setup.orientation = tab.ORIENTATION_LANDSCAPE
                tab.page_setup.paperSize = tab.PAPERSIZE_A3
                tab.title = tab_name
            else:
                tab = wb.create_sheet(title=tab_name)  # Excel allows max 31 characters in tab names

            self.wb_tabs_open[tab_id] = tab

            tab.sheet_properties.tabColor = tab_color
            tab.sheet_view.showGridLines = showGridLines

        self.sys_tab_seq = live_sys_tab_seq

        return wb

    def save_workbook(self, wb):
        logger.debug(f"Saving Spreadsheet: {self.sys_pn_wbs}")

        # save and load workbook
        if os.path.isfile(self.sys_pn_wbs):
            try_again = True
            while try_again:
                try:
                    os.remove(self.sys_pn_wbs)
                    try_again = False
                except PermissionError:
                    msg = f"Unable to save workbook {self.sys_pn_wbs}."
                    logger.warning(msg)
                    try_again = self.retry_file_removal(msg)
                    if not try_again:
                        msg += " User Canceled on Retry."
                        logger.critical(msg)
                        raise SystemExit(msg)

                except Exception as e:
                    raise Exception(f"Error removing file {self.sys_pn_wbs}: {e}")

        wb.save(self.sys_pn_wbs)

    def retry_file_removal(self, msg):
        # returns True on Retry
        result = messagebox.askretrycancel("Warning! File in use", msg)
        return result

    def obs_hyperlink(self, file):
    # vault can be either the vault name, or the vault ID.
    # The vault name is simply the name of the vault folder.
    # The vault ID is the random 16-character code assigned to the vault.
        # This ID is unique per folder on your computer. Example: ef6ca3e3b524d22f.
        # There isn't an easy way to find this ID yet, one will be offered at a
        # later date in the vault switcher. Currently it can be found in
        # %appdata%/obsidian/obsidian.json for Windows.
        # For MacOS, replace
        #   %appdata% with ~/Library/Application Support/.
        # For Linux, replace
        #   %appdata% with ~/.config/.
        file_link = f"{urllib.parse.quote(file, safe=':/')}"
        obs_link_text = file.replace(".md", "")
        obs_link = f'=hyperlink("obsidian://open?vault={self.vault_id}&file={file_link}","{obs_link_text}")'

        return obs_link

    @staticmethod
    def web_hyperlink(file):
        file_link = f"{urllib.parse.quote(file, safe=':/')}"
        web_link_text = file
        web_link = f'=hyperlink("{file_link}","{web_link_text}")'

        return web_link

    def format_as_table(self, tab, tbl_nm, tab_tbl_style, tot_rows):
        tbl_beg_col   = self.tab_def['tbl_beg_col']
        tbl_end_col   = self.tab_def['tbl_end_col']
        tbl_hdr_row   = self.tab_def['tbl_hdr_row']
        tbl_rng = f"{self.xl_a_col(tbl_beg_col)}{tbl_hdr_row}:{self.xl_a_col(tbl_end_col)}"
        if tot_rows == int((tot_rows - tbl_hdr_row)):
            tbl_rng = f"{tbl_rng}11"
        else:
            tbl_rng = tbl_rng + str(tot_rows - 1)

        logger.debug(f"tbl_name: {tbl_nm}  tbl_rng: {tbl_rng}")

        tbl = Table(displayName=tbl_nm, ref=tbl_rng)
        tbl_style = TableStyleInfo(name=tab_tbl_style, showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = tbl_style
        tab.add_table(tbl)

    def xl_clean_cell(self, cell_value):
        """
        Replace illegal XML characters in the given cell value with '@'
        if the cell value is a string. Handle datetime objects correctly.
        """
        # Check and process datetime objects directly
        if isinstance(cell_value, datetime):
            cell_value = cell_value.replace(tzinfo=None)

        # Process strings for illegal characters
        if isinstance(cell_value, str):
            # Strip TZ from dates in string format if applicable
            if self.rgx_noTZdatePattern.search(cell_value):
                cell_value = self.rgx_noTZdatePattern.sub(self.rgx_noTZdateReplace, cell_value)
            return ILLEGAL_CHARACTERS_RE.sub("z", cell_value)

        return cell_value

    @staticmethod
    def xl_a_col(col_num):
        col_alpha = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            col_alpha = chr(65 + remainder) + col_alpha
        return col_alpha

    @staticmethod
    def xl_set_border(ws: openpyxl.Workbook, b_parms: tuple) -> None:
        """
        Sets borders for a specified cell range in a worksheet with customizable options
        for border type, color, and sides.

        This function applies borders to cells within a given range in an Excel
        worksheet. Accepts attributes such as the type of border (e.g., thin,
        thick), the color of the border, and the sides of the cell to which the border
        is applied (ie, top, bottom, left, right, h-sides, v-sides, or all).

        If no values are provided for `border_type`, `color`, or `sides`, default
        values are used. The function allows setting borders to all sides of a cell or
        specific sides such as top, bottom, left, or right, including combinations like
        vertical sides or horizontal sides.

        :param ws: The worksheet object to modify.
        :type ws: openpyxl.worksheet.worksheet.Worksheet
        :param b_parms: A tuple containing settings for the border including the cell
            range, border type, color, and the side(s) for which to apply the border.
            The tuple should contain four elements in the following order:
            1. cell_range (str): The range of cells where borders are to be applied,
               specified in Excel range format (e.g., "A1:C3").
            2. border_type (str): The style of the border (e.g., "thin", "thick").
               Defaults to "thin" if not provided.
            3. color (str): The color of the border in hex color format without "#"
               (e.g., "000000" for black). Defaults to "000000" if not provided.
            4. sides (str): The side(s) of cells to apply the border on. Acceptable
               values are "all", "top", "bottom", "left", "right", "v-sides", or
               "h-sides". Defaults to "all" if not provided.
        :type b_parms: tuple
        :return: None
        :rtype: NoneType
        """

        cell_range, border_type, color, sides = b_parms
        if border_type is None or border_type == "":
            border_type = "thin"
        if color is None or color == "":
            color = "000000"
        if sides is None or sides == "":
            sides = "all"

        sides = sides.lower()
        border = Side(border_style=border_type, color=color)

        if sides == "all":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(top=border, bottom=border)
                    if border_first_col:
                        cell.border = Border(top=border, bottom=border, left=border)
                        border_first_col = False
                    if row[-1] == cell:
                        cell.border = Border(top=border, bottom=border, right=border)
        elif sides == "bottom":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(bottom=border)
        elif sides == "top":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(top=border)
        elif sides == "left":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(left=border)
        elif sides == "right":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(right=border)
        elif sides == "v-sides":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(left=border, right=border)
        elif sides == "h-sides":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(top=border, bottom=border)

def main() -> None:
    exporter = ExcelExporter()
    exporter.export()

if __name__ == '__main__':
    main()


