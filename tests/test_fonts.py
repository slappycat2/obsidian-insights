"""Tests for workbook font selection.

Issue #5: tab titles and subtitles asked for "Berlin Sans FB Demi" and
"Berlin Sans", which ship with Microsoft Office and exist on almost no macOS or
Linux machine. An .xlsx cell names exactly one font and has no fallback list,
so the request was silently substituted by whatever opened the file.
"""

from pathlib import Path

import openpyxl
import pytest

from vault_check import v_chk_wb_tabs as tabs

SOURCE = (Path(__file__).resolve().parent.parent
          / "src" / "vault_check" / "v_chk_wb_tabs.py").read_text(encoding="utf-8")


@pytest.fixture(autouse=True)
def clear_font_caches():
    """Both lookups are lru_cached, so a monkeypatched search must start clean."""
    tabs.font_is_installed.cache_clear()
    tabs.display_font.cache_clear()
    yield
    tabs.font_is_installed.cache_clear()
    tabs.display_font.cache_clear()


# ---------------------------------------------------------------------------
# Detection
# ---------------------------------------------------------------------------

def test_an_installed_font_is_found(tmp_path, monkeypatch):
    (tmp_path / "impact.ttf").write_bytes(b"")
    monkeypatch.setattr(tabs, "font_search_dirs", lambda: [tmp_path])

    assert tabs.font_is_installed("Impact")


def test_detection_looks_in_subdirectories(tmp_path, monkeypatch):
    """Linux nests fonts, e.g. /usr/share/fonts/truetype/msttcorefonts/."""
    nested = tmp_path / "truetype" / "msttcorefonts"
    nested.mkdir(parents=True)
    (nested / "Impact.ttf").write_bytes(b"")
    monkeypatch.setattr(tabs, "font_search_dirs", lambda: [tmp_path])

    assert tabs.font_is_installed("Impact")


def test_a_missing_font_is_not_found(tmp_path, monkeypatch):
    (tmp_path / "arial.ttf").write_bytes(b"")
    monkeypatch.setattr(tabs, "font_search_dirs", lambda: [tmp_path])

    assert not tabs.font_is_installed("Impact")


def test_a_partial_filename_match_does_not_count(tmp_path, monkeypatch):
    """"impactful.ttf" is not Impact. The match is on the whole stem."""
    (tmp_path / "impactful.ttf").write_bytes(b"")
    monkeypatch.setattr(tabs, "font_search_dirs", lambda: [tmp_path])

    assert not tabs.font_is_installed("Impact")


def test_missing_font_directories_are_survivable(tmp_path, monkeypatch):
    """Every listed directory is optional -- most machines have only some."""
    monkeypatch.setattr(tabs, "font_search_dirs", lambda: [tmp_path / "nope"])

    assert not tabs.font_is_installed("Impact")


def test_font_search_dirs_are_absolute_paths():
    for directory in tabs.font_search_dirs():
        assert directory.is_absolute(), directory


# ---------------------------------------------------------------------------
# Fallback
# ---------------------------------------------------------------------------

def test_display_font_is_the_display_font_when_installed(monkeypatch):
    monkeypatch.setattr(tabs, "font_is_installed", lambda family: True)

    assert tabs.display_font() == tabs.DISPLAY_FONT


def test_display_font_is_none_when_the_font_is_missing(monkeypatch):
    """None means "state no preference", which sends the cell down
    export_cell's fallback to FALLBACK_FONT."""
    monkeypatch.setattr(tabs, "font_is_installed", lambda family: False)

    assert tabs.display_font() is None


def test_no_tab_hardcodes_a_font_name():
    """Regression, issue #5: the two Berlin Sans names appeared 38 times across
    the tab subclasses, so the font could not be changed in one place."""
    body = SOURCE.partition("TITLE_FONT = display_font()")[2]

    assert "Berlin Sans" not in body
    assert "'Impact'" not in body and '"Impact"' not in body


# ---------------------------------------------------------------------------
# What actually reaches the workbook
# ---------------------------------------------------------------------------

@pytest.fixture
def workbook(make_vault, stub_config):
    from vault_check.v_chk_build import VaultHealthCheck
    from vault_check.v_chk_wb_tabs import NewWb
    from vault_check.v_chk_xl import ExcelExporter

    vault = make_vault({"Note.md": "---\nauthor: Jane\n---\n\nBody.\n"})
    exporter = ExcelExporter(NewWb(VaultHealthCheck(stub_config(vault))).wbd_obj)
    exporter.export()
    return openpyxl.load_workbook(exporter.sys_pn_wbs)


def test_tab_titles_use_the_resolved_display_font(workbook):
    """Whatever the machine running the tests has: Impact where it is
    installed, and the workbook default where it is not."""
    title = workbook["Properties"].cell(row=2, column=3)

    assert title.value == "Properties Analysis"
    assert title.font.name == tabs.TITLE_FONT


def test_no_cell_carries_an_empty_font_name(workbook):
    """A cell definition passing '' means "no preference". export_cell turns
    that into FALLBACK_FONT, so no cell reaches the file with <name val=""/>,
    which is an explicitly empty font name rather than an absent one."""
    empties = [f"{ws.title}!{cell.coordinate}"
               for ws in workbook.worksheets
               for row in ws.iter_rows()
               for cell in row
               if cell.font is not None and cell.font.name == ""]

    assert not empties, f"cells with an empty font name: {empties[:5]}"


def test_titles_name_no_font_at_all_when_the_display_font_is_missing(
        make_vault, stub_config, monkeypatch):
    """The graceful part of "degrade gracefully": on a machine without the
    display font, a title falls through to the same fallback every other
    unstyled cell uses, rather than naming a font that is not there and letting
    the viewer guess.

    The tab subclasses read TITLE_FONT when they are instantiated, so patching
    the module global here is enough.
    """
    from vault_check.v_chk_xl import FALLBACK_FONT
    from vault_check.v_chk_build import VaultHealthCheck
    from vault_check.v_chk_wb_tabs import NewWb
    from vault_check.v_chk_xl import ExcelExporter

    monkeypatch.setattr(tabs, "TITLE_FONT", None)

    vault = make_vault({"Note.md": "---\nauthor: Jane\n---\n\nBody.\n"})
    exporter = ExcelExporter(NewWb(VaultHealthCheck(stub_config(vault))).wbd_obj)
    exporter.export()

    title = openpyxl.load_workbook(exporter.sys_pn_wbs)["Properties"].cell(row=2, column=3)

    assert title.value == "Properties Analysis"
    assert title.font.name == FALLBACK_FONT
