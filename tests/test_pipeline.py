"""End-to-end test: markdown in, .xlsx out.

Covers what the parsing tests cannot -- the YAML batch-file handoff between
stages, the tab definitions, and the openpyxl export. No GUI and no Obsidian
installation are involved.
"""

import openpyxl
import pytest

VAULT = {
    "Projects/Alpha.md": """
        ---
        author: Jane
        status: active
        tags:
          - project
          - alpha
        ---

        Alpha body. rating:: 5

        ```dataview
        TABLE file.name
        ```
    """,
    "Projects/Beta.md": """
        ---
        author: Sam
        status: done
        ---

        Beta body with an inline #beta tag.
    """,
    "Archive/Alpha.md": """
        ---
        author: Jane
        ---

        A duplicate filename, in a different folder.
    """,
    # One note supplying two values of the same property: two values, but one
    # file. This is the case the Properties tab's file count used to double.
    "Projects/Gamma.md": """
        ---
        contributors:
          - Ann
          - Bob
        ---

        Gamma body.
    """,
    "Broken.md": """
        ---
        author: [unclosed
        ---

        Body.
    """,
}


@pytest.fixture
def workbook_path(make_vault, stub_config):
    """Run all four stages over a small vault and return the .xlsx path."""
    from vault_check.v_chk_build import VaultHealthCheck
    from vault_check.v_chk_wb_tabs import NewWb
    from vault_check.v_chk_xl import ExcelExporter

    vault = make_vault(VAULT)

    health_check = VaultHealthCheck(stub_config(vault))
    vault_wb = NewWb(health_check)

    exporter = ExcelExporter(vault_wb.wbd_obj)
    exporter.export()

    return exporter.sys_pn_wbs


def test_workbook_is_created(workbook_path):
    from pathlib import Path

    assert Path(workbook_path).is_file()
    assert Path(workbook_path).stat().st_size > 0


def test_workbook_has_the_expected_tabs(workbook_path):
    wb = openpyxl.load_workbook(workbook_path)

    # Tabs whose data source is empty are deliberately dropped, so assert on a
    # subset that this vault definitely populates.
    for expected in ("Summary", "Properties", "Values", "Tags", "Files"):
        assert expected in wb.sheetnames


def test_properties_reach_the_workbook(workbook_path):
    wb = openpyxl.load_workbook(workbook_path)
    text = {
        str(cell.value)
        for row in wb["Values"].iter_rows()
        for cell in row
        if cell.value is not None
    }

    assert "author" in text
    assert "Jane" in text


def _properties_rows(workbook_path):
    """The Properties tab as {property: (values count, file count)}.

    Columns are RowId 10, Properties 11, Values 12, Links 13; the header is on
    row 10 and the data starts at 11.
    """
    wb = openpyxl.load_workbook(workbook_path)
    tab = wb["Properties"]

    rows = {}
    for row in range(11, tab.max_row + 1):
        name = tab.cell(row=row, column=11).value
        if name is None or name == "":
            break
        rows[str(name)] = (tab.cell(row=row, column=12).value,
                           tab.cell(row=row, column=13).value)
    return rows


def test_properties_tab_headers(workbook_path):
    """The third column counts distinct notes, so it is "Files", not "Links".

    The name is also the Excel table column name, which the Summary tab's
    tbl_pros[...] formulas reference -- they derive from the same string, so a
    mismatch here would break those formulas rather than just the label.
    """
    wb = openpyxl.load_workbook(workbook_path)
    tab = wb["Properties"]

    headers = [tab.cell(row=10, column=c).value for c in range(10, 14)]

    assert headers == ["RowId", "Properties", "Values", "Files"]


def test_every_property_gets_exactly_one_row(workbook_path):
    """Regression, issue #1: rows were emitted at the boundary between two
    properties, so the tab opened with a phantom row carrying no property name
    and the last property in the vault was never written at all."""
    rows = _properties_rows(workbook_path)

    assert set(rows) == {"author", "status", "contributors"}
    assert "Properties" not in rows, "phantom row named after the column header"


def test_property_file_count_is_distinct_files(workbook_path):
    """Regression, issue #1: the file count summed each value's file list, so a
    note using two values of one property was counted twice.

    'contributors' is the case that matters -- Projects/Gamma.md supplies both
    values, so it is two values but one file.
    """
    rows = _properties_rows(workbook_path)

    assert rows["contributors"] == (2, 1)


def test_property_row_counts_match_the_vault(workbook_path):
    """The other side of the same regression: 'author' has two values spread
    over three files, and used to report two."""
    rows = _properties_rows(workbook_path)

    assert rows["author"] == (2, 3)
    assert rows["status"] == (2, 2)


def test_summary_tab_shows_the_running_version(workbook_path):
    """Regression: the Summary tab used to announce two hardcoded versions --
    'v1.0' in the title and 'v.0.9 (beta)' in the hyperlink -- while the package
    said 0.3.0 and CONFIG.yaml said 0.2.9. This asserts the rendered workbook,
    which is the only place a user ever sees the number."""
    from vault_check import __version__

    wb = openpyxl.load_workbook(workbook_path)
    text = " ".join(
        str(cell.value)
        for row in wb["Summary"].iter_rows()
        for cell in row
        if cell.value is not None
    )

    assert f"v{__version__}" in text
    assert "v1.0" not in text
    assert "0.9 (beta)" not in text


def test_duplicate_filenames_reach_the_workbook(workbook_path):
    wb = openpyxl.load_workbook(workbook_path)
    assert "Duplicates" in wb.sheetnames

    text = " ".join(
        str(cell.value)
        for row in wb["Duplicates"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Alpha.md" in text


def test_broken_yaml_reaches_the_workbook(workbook_path):
    wb = openpyxl.load_workbook(workbook_path)
    assert "Xyml" in wb.sheetnames

    text = " ".join(
        str(cell.value)
        for row in wb["Xyml"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Broken.md" in text


def test_tables_are_defined_on_data_tabs(workbook_path):
    """The workbook's value depends on real Excel tables (filters and the
    AGGREGATE-based totals), not just cell values."""
    wb = openpyxl.load_workbook(workbook_path)

    assert "tbl_pros" in wb["Properties"].tables
    assert "tbl_vals" in wb["Values"].tables


def test_templates_tab_reaches_the_workbook(make_vault, stub_config):
    """Regression, issue #6: nothing populated obs_tmplt, so the Templates tab
    was always empty -- and initialize_all_tabs() drops empty tabs, so it never
    reached a workbook at all.

    Asserted end to end because the drop happens during export, not harvesting.
    """
    from vault_check.v_chk_build import VaultHealthCheck
    from vault_check.v_chk_wb_tabs import NewWb
    from vault_check.v_chk_xl import ExcelExporter

    vault = make_vault({
        "Notes/Real.md": "---\nauthor: Jane\n---\n\nReal note.\n",
        "Templates/Daily.md": (
            '---\nauthor: PLACEHOLDER\ndate: <% tp.date.now("YYYY-MM-DD") %>\n'
            "tags: [daily]\n---\n\n# <% tp.file.title %>\n"
        ),
    })
    health_check = VaultHealthCheck(stub_config(vault, dir_templates=str(vault / "Templates")))
    exporter = ExcelExporter(NewWb(health_check).wbd_obj)
    exporter.export()

    wb = openpyxl.load_workbook(exporter.sys_pn_wbs)
    assert "Templates" in wb.sheetnames

    # This tab's table starts at a different column from the Properties tab:
    # RowId 4, Property 5, Values 6. Pinned by the header assertion below.
    tab = wb["Templates"]
    assert [tab.cell(row=10, column=c).value for c in range(4, 7)] == \
           ["RowId", "Property", "Values"]

    rows = {tab.cell(row=r, column=5).value: tab.cell(row=r, column=6).value
            for r in range(11, tab.max_row + 1)}

    assert rows.get("author") == "PLACEHOLDER"
    assert rows.get("tags") == "daily"
    # `date: <% tp.date.now() %>` loses its value when Templater tags are
    # stripped, and is marked rather than left blank.
    assert rows.get("date") == "(-None-)"
